import os
from openpyxl import load_workbook, Workbook

def read_cell_from_excel(file_path, cell):
    """
    Opens an Excel file and reads a specific cell.
    """
    workbook = load_workbook(filename=file_path, data_only=True)
    first_sheet_name = workbook.sheetnames[0]  # Assumes data is in the first sheet
    sheet = workbook[first_sheet_name]
    return sheet[cell].value

def write_results_to_excel(output_file, results):
    """
    Writes results to an Excel file. Results should be a list of tuples (filename, cell_value).
    """
    workbook = Workbook()
    sheet = workbook.active
    
    # Optionally, set headers
    sheet.append(["Filename", "Cell Value"])
    
    for filename, cell_value in results:
        sheet.append([filename, cell_value])
    
    workbook.save(filename=output_file)

data_folder = 'data_folder'
result_folder = 'result_folder'
result_file = 'summary.xlsx'  # Name of the result file in result_folder

# Ensure the result folder exists
if not os.path.exists(result_folder):
    os.makedirs(result_folder)

# Path for the result file
result_file_path = os.path.join(result_folder, result_file)

# List to store tuples of (filename, cell content)
results = []

# Loop through each file in data_folder
for filename in os.listdir(data_folder):
    if filename.endswith('.xlsx'):  # Ensure we're only processing Excel files
        file_path = os.path.join(data_folder, filename)
        
        # Read cell C9 from the current file
        cell_value = read_cell_from_excel(file_path, 'C9')
        
        # Append the filename and cell value to the results list
        results.append((filename, cell_value))

# Write the results to the excel file in result_folder
write_results_to_excel(result_file_path, results)

print(f"Process completed. Results have been written to {result_file_path}")
